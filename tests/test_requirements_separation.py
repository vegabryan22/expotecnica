import json
import io
import unittest
from pathlib import Path
from unittest.mock import patch

from sqlalchemy import create_engine, text

from app import _reconcile_existing_logistics_statuses, create_app, natural_title
from app.controllers.admin_controller import (
    ACTION_MODULE_MAP,
    ADMIN_DEPARTMENT_MODULE_ACCESS,
    _build_logistics_pending_report_rows,
    _build_project_logistics_summary,
    _build_tutor_logistics_reminder_payload,
    _build_tutor_logistics_digest_payload,
    _build_logistics_reminder_data,
    _build_exposition_usher_report_rows,
    _judge_report_rows,
    _build_advisor_stats,
    _person_name_title,
    _mysql_base_args,
    _project_logistics_progress,
    _project_report_rows,
    _sync_project_logistics_status,
    _sync_project_photo_validation,
)
from app.controllers.project_controller import _build_requirement_items, _get_or_create_tutor_atomic
from app.extensions import db
from app.models.assignment import Assignment
from app.models.judge import Judge
from app.models.project import Project
from app.models.project_member import ProjectMember
from app.models.tutor import Tutor
from app.models.venue import Venue
from app.services import mail_service


class RequirementsSeparationTest(unittest.TestCase):

    def test_project_cards_show_saved_observations(self):
        template = Path("app/templates/admin/projects.html").read_text(encoding="utf-8")

        self.assertIn("project.logistics_notes", template)
        self.assertIn("project.requirements_notes", template)
        self.assertIn("project-observations", template)
        self.assertIn("Observaciones", template)

    def test_mysql_cli_commands_force_tcp_even_for_localhost(self):
        config = {
            "host": "localhost",
            "port": "3306",
            "user": "expotecnica",
        }

        args = _mysql_base_args("mysqldump", config)

        self.assertIn("--protocol=TCP", args)
        self.assertEqual("localhost", args[args.index("--host") + 1])
        self.assertEqual("3306", args[args.index("--port") + 1])

    def test_student_attention_is_an_operational_point_not_a_project_venue(self):
        legacy_venue = Venue(name="Atención Estudiantes", venue_type="otro")
        explicit_venue = Venue(name="Servicio al estudiante", venue_type=Venue.TYPE_STUDENT_ATTENTION)

        self.assertFalse(legacy_venue.accepts_projects)
        self.assertFalse(explicit_venue.accepts_projects)
        self.assertEqual("Punto de atención estudiantil", legacy_venue.operational_label)
        self.assertEqual("Atención estudiantes", explicit_venue.type_label)

    def test_superadmin_can_impersonate_and_return_without_target_password(self):
        app = create_app()
        app.config["TESTING"] = True
        with app.app_context():
            admin = Judge.query.filter(Judge.role == Judge.ROLE_SUPERADMIN).first()
            self.assertIsNotNone(admin)
            target = Judge(
                full_name="Usuario temporal para soporte",
                email="impersonation.test@example.invalid",
                role=Judge.ROLE_USHER_LOGISTICS,
                is_admin=False,
                is_active_user=True,
            )
            target.set_password("ClaveQueNoSeUsa123")
            db.session.add(target)
            db.session.commit()
            target_id = target.id
            admin_id = admin.id
            try:
                with app.test_client() as client:
                    with client.session_transaction() as user_session:
                        user_session["_user_id"] = str(admin_id)
                        user_session["_fresh"] = True

                    start = client.post(f"/admin/usuarios/{target_id}/suplantar")
                    self.assertEqual(302, start.status_code)
                    self.assertIn("/admin/logistica-edecanes", start.headers.get("Location", ""))
                    with client.session_transaction() as user_session:
                        self.assertEqual(str(target_id), user_session.get("_user_id"))
                        self.assertEqual(admin_id, user_session.get("impersonator_user_id"))

                    panel = client.get("/admin/logistica-edecanes")
                    self.assertIn("Modo soporte activo", panel.get_data(as_text=True))
                    self.assertIn("Volver a mi cuenta", panel.get_data(as_text=True))

                    stop = client.post("/auth/finalizar-soporte")
                    self.assertEqual(302, stop.status_code)
                    self.assertIn("/admin/jueces", stop.headers.get("Location", ""))
                    with client.session_transaction() as user_session:
                        self.assertEqual(str(admin_id), user_session.get("_user_id"))
                        self.assertNotIn("impersonator_user_id", user_session)
            finally:
                db.session.delete(Judge.query.get(target_id))
                db.session.commit()

    def test_non_superadmin_cannot_impersonate_a_user(self):
        app = create_app()
        app.config["TESTING"] = True
        with app.app_context():
            regular_admin = Judge.query.filter(Judge.role == Judge.ROLE_ADMIN).first()
            target = Judge.query.filter(Judge.id != regular_admin.id).first() if regular_admin else None
            if regular_admin is None or target is None:
                self.skipTest("Se requieren un administrador y otro usuario en los datos de prueba")
            with app.test_client() as client:
                with client.session_transaction() as user_session:
                    user_session["_user_id"] = str(regular_admin.id)
                    user_session["_fresh"] = True
                response = client.post(f"/admin/usuarios/{target.id}/suplantar")
            self.assertEqual(403, response.status_code)

    def test_usher_logistics_role_has_read_only_operational_access(self):
        app = create_app()
        app.config["TESTING"] = True
        with app.app_context():
            user = Judge(
                full_name="Prueba Logística Edecanes",
                email="prueba.edecanes.role@example.invalid",
                role=Judge.ROLE_USHER_LOGISTICS,
                is_admin=False,
                is_active_user=True,
                can_evaluate_documentation=False,
                can_evaluate_exposition=False,
            )
            user.set_password("Temporal123")
            db.session.add(user)
            db.session.commit()
            try:
                with app.test_client() as client:
                    with client.session_transaction() as session:
                        session["_user_id"] = str(user.id)
                        session["_fresh"] = True
                    hub = client.get("/admin/logistica-edecanes")
                    map_pdf = client.get("/admin/recintos/mapa/imprimir")
                    forbidden = client.get("/admin/jueces")
                self.assertEqual(200, hub.status_code)
                self.assertIn("Logística de edecanes", hub.get_data(as_text=True))
                self.assertEqual(200, map_pdf.status_code)
                self.assertEqual(302, forbidden.status_code)
                self.assertIn("/admin/logistica-edecanes", forbidden.headers.get("Location", ""))
            finally:
                db.session.delete(user)
                db.session.commit()

    def test_exposition_capacity_draft_renders_without_applying_changes(self):
        app = create_app()
        app.config["TESTING"] = True
        with app.app_context():
            admin = Judge.query.filter(Judge.role == Judge.ROLE_SUPERADMIN).first()
            self.assertIsNotNone(admin)
            before = [
                (row.id, row.judge_id, row.project_id, row.can_evaluate_documentation, row.can_evaluate_exposition)
                for row in Assignment.query.order_by(Assignment.id.asc()).all()
            ]
            with app.test_client() as client:
                with client.session_transaction() as session:
                    session["_user_id"] = str(admin.id)
                    session["_fresh"] = True
                response = client.get("/admin/asignaciones/cupo-presencial")
            after = [
                (row.id, row.judge_id, row.project_id, row.can_evaluate_documentation, row.can_evaluate_exposition)
                for row in Assignment.query.order_by(Assignment.id.asc()).all()
            ]

        self.assertEqual(200, response.status_code)
        self.assertIn("Planificador de cupo presencial", response.get_data(as_text=True))
        self.assertEqual(before, after)

    def test_detce_forms_report_renders_calculated_answers(self):
        app = create_app()
        app.config["TESTING"] = True
        with app.app_context():
            admin = Judge.query.filter(Judge.role == Judge.ROLE_SUPERADMIN).first()
            self.assertIsNotNone(admin)
            with app.test_client() as client:
                with client.session_transaction() as session:
                    session["_user_id"] = str(admin.id)
                    session["_fresh"] = True
                response = client.get("/admin/reportes/respuestas-detce")

        self.assertEqual(200, response.status_code)
        body = response.get_data(as_text=True)
        self.assertIn("Formulario DETCE", body)
        self.assertIn("Cantidad total de proyectos", body)

    def test_reports_center_renders_for_admin(self):
        app = create_app()
        app.config["TESTING"] = True

        with app.app_context():
            admin = Judge.query.filter(Judge.role == Judge.ROLE_SUPERADMIN).first()
            self.assertIsNotNone(admin)

            with app.test_client() as client:
                with client.session_transaction() as session:
                    session["_user_id"] = str(admin.id)
                    session["_fresh"] = True

                response = client.get("/admin/reportes")

        self.assertEqual(200, response.status_code)
        self.assertIn("Reportes de ExpoTécnica", response.get_data(as_text=True))

    def test_venues_maintenance_and_usher_workbook_are_available(self):
        from openpyxl import load_workbook
        from pypdf import PdfReader

        app = create_app()
        app.config["TESTING"] = True
        with app.app_context():
            admin = Judge.query.filter(Judge.role == Judge.ROLE_SUPERADMIN).first()
            self.assertIsNotNone(admin)
            with app.test_client() as client:
                with client.session_transaction() as session:
                    session["_user_id"] = str(admin.id)
                    session["_fresh"] = True
                venues_response = client.get("/admin/recintos")
                printable_map = client.get("/admin/recintos/mapa/imprimir")
                report_response = client.get("/admin/asignaciones/reporte/edecanes/excel")

        self.assertEqual(200, venues_response.status_code)
        self.assertIn("Recintos", venues_response.get_data(as_text=True))
        self.assertEqual(200, printable_map.status_code)
        self.assertEqual("application/pdf", printable_map.mimetype)
        self.assertTrue(printable_map.data.startswith(b"%PDF-"))
        self.assertGreater(len(printable_map.data), 100_000)
        map_reader = PdfReader(io.BytesIO(printable_map.data))
        self.assertGreaterEqual(len(map_reader.pages), 1)
        first_page_text = map_reader.pages[0].extract_text() or ""
        self.assertIn("PUNTOS OPERATIVOS", first_page_text)
        if len(map_reader.pages) > 1:
            directory_text = "\n".join(page.extract_text() or "" for page in map_reader.pages[1:])
            self.assertIn("Directorio completo de proyectos por recinto", directory_text)
        self.assertEqual(200, report_response.status_code)
        workbook = load_workbook(io.BytesIO(report_response.data), read_only=True)
        self.assertEqual(["Jueces", "Integrantes"], workbook.sheetnames)
        self.assertEqual(["Juez", "Proyecto", "Recinto"], [cell.value for cell in next(workbook["Jueces"].iter_rows(min_row=5, max_row=5))])
        self.assertEqual(["Integrante", "Proyecto", "Recinto"], [cell.value for cell in next(workbook["Integrantes"].iter_rows(min_row=1, max_row=1))])

    def test_gmail_requires_account_and_app_password(self):
        values = {
            "smtp_provider": "gmail",
            "smtp_host": "smtp.gmail.com",
            "smtp_port": "587",
            "smtp_username": "expotec@example.com",
            "smtp_password": "app-password",
            "smtp_from_email": "expotec@example.com",
            "smtp_use_tls": "1",
            "smtp_use_ssl": "0",
        }
        with patch.object(
            mail_service.SystemSetting,
            "get_value",
            side_effect=lambda key, default=None: values.get(key, default),
        ):
            self.assertTrue(mail_service.smtp_is_configured())
            values["smtp_password"] = ""
            self.assertFalse(mail_service.smtp_is_configured())

    def test_smtp_page_offers_guided_gmail_setup(self):
        template = Path("app/templates/admin/smtp.html").read_text(encoding="utf-8")

        self.assertIn('value="gmail"', template)
        self.assertIn("smtp.gmail.com", template)
        self.assertIn("contrase&ntilde;a de aplicaci&oacute;n", template)
        self.assertIn("myaccount.google.com/apppasswords", template)
        self.assertIn("{{ 'Sí' if smtp_settings.use_tls else 'No' }}", template)
        self.assertNotIn("{{ 'S&iacute;' if smtp_settings.use_tls else 'No' }}", template)
    def test_tutor_statistics_centralize_projects_students_and_pending_work(self):
        complete = Project(
            id=31,
            title="Proyecto completo",
            team_name="Equipo A",
            advisor_name="MARÍA ELENA DEL RÍO",
            advisor_identity="123456789",
            advisor_email="tutora@example.com",
            advisor_phone="88880000",
            advisor_specialty="Informática",
            category="steam",
            is_active=True,
            logistics_status="completo",
            requirements_status="completo",
            project_document_path="document.pdf",
            project_logo_path="logo.png",
            logistics_document_ok=True,
            logistics_logo_ok=True,
            logistics_photos_ok=True,
            logistics_registration_form_signed_ok=True,
            logistics_student_consents_signed_ok=True,
            logistics_cedula_tutor_ok=True,
        )
        complete.members = [
            ProjectMember(full_name="Estudiante", student_number=1, section_name="12-1", photo_url="student.jpg", consent_signed_ok=True, cedula_encargado_ok=True, cedula_estudiante_ok=True)
        ]
        pending = Project(
            id=32,
            title="Proyecto pendiente",
            team_name="Equipo B",
            advisor_name="María Elena del Río",
            advisor_identity="123456789",
            advisor_email="tutora@example.com",
            category="emprendimiento",
            is_active=True,
            logistics_status="incompleto",
            requirements_status="pendiente_revision",
        )
        pending.members = [ProjectMember(full_name="Otro", student_number=1, section_name="11-2")]

        tutors = _build_advisor_stats([complete, pending])

        self.assertEqual(1, len(tutors))
        self.assertEqual("María Elena del Río", tutors[0]["name"])
        self.assertEqual(2, tutors[0]["total"])
        self.assertEqual(2, tutors[0]["students"])
        self.assertEqual(1, tutors[0]["completed"])
        self.assertEqual(1, tutors[0]["pending"])
        self.assertEqual("11-2, 12-1", tutors[0]["sections_label"])
        self.assertEqual(1, tutors[0]["pending_total"])

    def test_tutor_logistics_percentage_reflects_partial_progress(self):
        project = Project(
            id=90,
            title="Avance parcial",
            team_name="Equipo",
            advisor_name="Tutor Ejemplo",
            advisor_identity="900000001",
            category="steam",
            is_active=True,
            logistics_status="incompleto",
            requirements_status="completo",
            project_document_path="proyecto.pdf",
            logistics_document_ok=True,
        )
        project.members = [ProjectMember(full_name="Estudiante", student_number=1)]

        completed, total = _project_logistics_progress(project)
        tutors = _build_advisor_stats([project])

        self.assertEqual((2, 10), (completed, total))
        self.assertEqual(20, tutors[0]["completion_percent"])

    def test_tutors_page_has_filters_statistics_and_reports_center(self):
        template = Path("app/templates/admin/tutors.html").read_text(encoding="utf-8")

        self.assertIn("tutors_summary", template)
        self.assertIn("tutors-filter-text", template)
        self.assertIn("reports_page", template)
        self.assertIn('name="action" value="update_advisor"', template)
        self.assertIn('name="action" value="toggle_tutor"', template)
        self.assertIn("tutors-export-btn", template)
        stylesheet = Path("app/static/style.css").read_text(encoding="utf-8")
        self.assertIn(".tutor-card-actions button", stylesheet)
        self.assertIn("grid-template-columns: repeat(5, minmax(0, 1fr));", stylesheet)

    def test_registration_uses_private_central_tutor_catalog(self):
        template = Path("app/templates/public/register_project.html").read_text(encoding="utf-8")

        self.assertIn('name="tutor_mode" value="existing"', template)
        self.assertIn('name="tutor_id"', template)
        self.assertIn("Registrar otro tutor", template)
        self.assertIn("La información privada permanece protegida", template)
        self.assertNotIn("tutor.identity_number", template)
        self.assertNotIn("tutor.birth_date", template)
        self.assertEqual("tutors", Tutor.__tablename__)
        self.assertIn("toggle_tutor", ACTION_MODULE_MAP)

    def test_tutor_creation_uses_atomic_mysql_upsert(self):
        import inspect

        source = inspect.getsource(_get_or_create_tutor_atomic)
        self.assertIn("ON DUPLICATE KEY UPDATE", source)
        self.assertIn("LAST_INSERT_ID(id)", source)

    def test_registration_only_requests_mentor_data_when_applicable(self):
        template = Path("app/templates/public/register_project.html").read_text(encoding="utf-8")

        self.assertIn('name="mentor_has" value="si"', template)
        self.assertIn('data-mentor-fields', template)
        self.assertIn("toggleMentorFields", template)
        self.assertIn('type="checkbox" name="declaration" value="si"', template)
        self.assertNotIn("No acepto", template)
        base_template = Path("app/templates/base.html").read_text(encoding="utf-8")
        self.assertIn("?v={{ asset_version }}", base_template)
        stylesheet = Path("app/static/style.css").read_text(encoding="utf-8")
        self.assertIn(".judge-register-band .form-grid.inline-grid-1", stylesheet)

    def test_member_photos_are_validated_automatically(self):
        project = Project(logistics_photos_ok=False)
        project.members = [
            ProjectMember(full_name="Estudiante uno", student_number=1, photo_url="uploads/uno.jpg"),
            ProjectMember(full_name="Estudiante dos", student_number=2, photo_url="uploads/dos.jpg"),
        ]

        self.assertTrue(_sync_project_photo_validation(project))
        self.assertTrue(project.logistics_photos_ok)

        project.members[1].photo_url = None

        self.assertFalse(_sync_project_photo_validation(project))
        self.assertFalse(project.logistics_photos_ok)

    def test_person_names_are_exported_with_natural_capitalization(self):
        self.assertEqual("María José de la Cruz", _person_name_title("MARÍA JOSÉ DE LA CRUZ"))
        self.assertEqual("Ana-María del Río", _person_name_title("ana-maría DEL RÍO"))
        self.assertEqual("Andrés Delgado Zúñiga", natural_title("ANDRÉS DELGADO ZÚÑIGA"))
        self.assertEqual(
            "Configuración y Soporte a Redes de Comunicación y Sistemas Operativos",
            natural_title("CONFIGURACIÓN Y SOPORTE A REDES DE COMUNICACIÓN Y SISTEMAS OPERATIVOS"),
        )

    def test_projects_report_contains_projects_and_members(self):
        project = Project(
            id=7,
            title="Proyecto de prueba",
            team_name="Equipo",
            representative_name="JUAN CARLOS DE LA O",
            representative_email="representante@example.com",
            advisor_name="MARÍA ELENA DEL RÍO",
            category="steam",
            description="Descripción",
            is_active=True,
            logistics_status="completo",
            requirements_status="completo",
        )
        project.members = [
            ProjectMember(
                student_number=1,
                full_name="ANA SOFÍA DE LOS ÁNGELES",
                section_name="12-1",
                specialty="Redes",
            )
        ]

        projects, members = _project_report_rows([project], {"steam": "STEAM"})

        self.assertEqual("Juan Carlos de la O", projects[0]["representative"])
        self.assertEqual("María Elena del Río", projects[0]["advisor"])
        self.assertEqual("Ana Sofía de los Ángeles", members[0]["name"])
        self.assertEqual("12-1", members[0]["section"])

    def test_projects_page_links_reports_center(self):
        template = Path("app/templates/admin/projects.html").read_text(encoding="utf-8")

        self.assertIn("reports_page", template)
        self.assertIn("Reportes centralizados", template)

    def test_projects_page_links_reminder_center(self):
        template = Path("app/templates/admin/projects.html").read_text(encoding="utf-8")

        self.assertIn("logistics_reminder_page", template)
        self.assertIn("Centro de recordatorios", template)

    def test_reports_center_centralizes_downloads(self):
        template = Path("app/templates/admin/reports.html").read_text(encoding="utf-8")
        controller = Path("app/controllers/admin_controller.py").read_text(encoding="utf-8")
        routes = Path("app/routes/admin_routes.py").read_text(encoding="utf-8")

        self.assertIn("/reportes", routes)
        self.assertIn("reports_page", controller)
        self.assertIn("Centro de reportes", template)
        self.assertIn("projects_report_excel", controller)
        self.assertIn("tutors_report_excel", controller)
        self.assertIn("judges_report_excel", controller)
        self.assertIn("exposition_usher_report_excel", controller)
        self.assertIn("participation_certificates_download", controller)

    def test_member_editor_can_delete_current_photo(self):
        template = Path("app/templates/admin/projects.html").read_text(encoding="utf-8")
        controller = Path("app/controllers/admin_controller.py").read_text(encoding="utf-8")

        self.assertIn('value="delete_member_photo"', template)
        self.assertIn("Eliminar foto actual", template)
        self.assertIn('"delete_member_photo": "projects"', controller)
        self.assertIn('member.photo_url = None', controller)
        self.assertIn('_sync_project_logistics_status(member.project)', controller)
        self.assertIn("Reportes centralizados", template)
        self.assertIn("Abrir reportes", template)

    def test_members_dialog_uses_profile_cards_and_collapsible_history(self):
        template = Path("app/templates/admin/projects.html").read_text(encoding="utf-8")

        self.assertIn('class="members-profile-grid"', template)
        self.assertIn('class="member-profile-card"', template)
        self.assertIn('class="members-history"', template)
        self.assertIn("Historial de cambios", template)
        self.assertNotIn("BitÃ¡cora de cambios de integrantes", template)
        self.assertIn('data-parent-dialog="members-project-{{ project.id }}"', template)
        self.assertIn('{{ next_url }}#members-project-{{ project.id }}', template)
        self.assertIn('class="danger-btn member-delete-btn"', template)

    def test_registration_builds_structured_requirement_items(self):
        form_data = {
            "requirement_item_name": ["Mesa de exhibiciÃ³n", "ExtensiÃ³n elÃ©ctrica"],
            "requirement_item_quantity": ["2", "1"],
            "requirement_item_unit": ["unidades", "unidad"],
            "requirement_item_notes": ["De 1,80 m", "De 10 metros"],
        }

        items = _build_requirement_items(form_data)

        self.assertEqual(2, len(items))
        self.assertEqual("Mesa de exhibiciÃ³n", items[0]["name"])
        self.assertEqual("2", items[0]["quantity"])
        self.assertEqual("unidades", items[0]["unit"])
        self.assertFalse(items[0]["confirmed"])

    def test_legacy_resource_text_is_preserved_for_detailing(self):
        project = Project(
            required_resources="Son tres estudiantes",
            requirements_resources_ok=False,
        )

        self.assertEqual(1, len(project.detailed_requirement_items))
        self.assertEqual("Son tres estudiantes", project.detailed_requirement_items[0]["name"])
        self.assertTrue(project.detailed_requirement_items[0]["legacy"])
        self.assertIn("pendiente de desglosar", project.detailed_requirement_items[0]["notes"])

    def test_requested_resources_have_their_own_completion_state(self):
        project = Project(
            requirements_summary="corriente, internet",
            requirements_items_json=json.dumps(
                [
                    {
                        "id": "item-1",
                        "name": "Mesa de exhibiciÃ³n",
                        "quantity": "2",
                        "unit": "unidades",
                        "notes": "De 1,80 m",
                        "confirmed": False,
                    }
                ]
            ),
            requirements_status="completo",
            requirements_current_ok=True,
            requirements_internet_ok=False,
        )

        self.assertEqual(
            project.requirements_missing_items,
            ["Acceso a internet", "Insumos pendientes: Mesa de exhibiciÃ³n"],
        )
        self.assertFalse(project.requirements_complete)

        project.requirements_internet_ok = True
        project.requirements_items_json = json.dumps(
            [
                {
                    "id": "item-1",
                    "name": "Mesa de exhibiciÃ³n",
                    "quantity": "2",
                    "unit": "unidades",
                    "notes": "De 1,80 m",
                    "confirmed": True,
                }
            ]
        )

        self.assertTrue(project.requirements_complete)

    def test_logistics_completion_does_not_depend_on_resource_requirements(self):
        project = Project(
            project_document_path="uploads/projects/document.pdf",
            project_logo_path="uploads/projects/logo.png",
            logistics_document_ok=True,
            logistics_logo_ok=True,
            logistics_photos_ok=True,
            logistics_registration_form_signed_ok=True,
            logistics_student_consents_signed_ok=True,
            logistics_cedula_tutor_ok=True,
            logistics_requirements_reviewed_ok=False,
        )

        self.assertTrue(project.logistics_requirements_complete)

    def test_logistics_status_is_completed_automatically(self):
        project = Project(
            project_document_path="uploads/projects/document.pdf",
            project_logo_path="uploads/projects/logo.png",
            logistics_document_ok=True,
            logistics_logo_ok=True,
            logistics_photos_ok=True,
            logistics_registration_form_signed_ok=True,
            logistics_student_consents_signed_ok=True,
            logistics_cedula_tutor_ok=True,
            logistics_status="pendiente_revision",
        )
        project.members = [
            ProjectMember(
                full_name="Estudiante listo",
                student_number=1,
                photo_url="uploads/student.jpg",
                consent_signed_ok=True,
                cedula_encargado_ok=True,
                cedula_estudiante_ok=True,
            )
        ]

        missing_items = _sync_project_logistics_status(project)

        self.assertEqual([], missing_items)
        self.assertEqual("completo", project.logistics_status)

        project.logistics_logo_ok = False
        missing_items = _sync_project_logistics_status(project)

        self.assertIn("logo validado", missing_items)
        self.assertEqual("incompleto", project.logistics_status)

    def test_logistics_status_stays_incomplete_while_identity_copies_are_missing(self):
        project = Project(
            project_document_path="uploads/projects/document.pdf",
            project_logo_path="uploads/projects/logo.png",
            logistics_document_ok=True,
            logistics_logo_ok=True,
            logistics_photos_ok=True,
            logistics_registration_form_signed_ok=True,
            logistics_student_consents_signed_ok=True,
            logistics_cedula_tutor_ok=False,
            logistics_status="completo",
        )
        project.members = [
            ProjectMember(
                full_name="Estudiante pendiente",
                student_number=1,
                photo_url="uploads/student.jpg",
                consent_signed_ok=True,
                cedula_encargado_ok=True,
                cedula_estudiante_ok=False,
            )
        ]

        self.assertEqual("incompleto", project.logistics_effective_status)
        missing_items = _sync_project_logistics_status(project)

        self.assertEqual("incompleto", project.logistics_status)
        self.assertIn("cedula del tutor", missing_items)
        self.assertTrue(any("cedula del estudiante" in item for item in missing_items))

    def test_tutor_project_badges_use_effective_logistics_status(self):
        template = Path("app/templates/admin/tutors.html").read_text(encoding="utf-8")

        self.assertIn("project.logistics_effective_status", template)
        self.assertNotIn("project.logistics_status == 'completo'", template)

    def test_existing_ready_projects_are_reconciled(self):
        engine = create_engine("sqlite://")
        with engine.begin() as connection:
            connection.execute(
                text(
                    """
                    CREATE TABLE projects (
                        id INTEGER PRIMARY KEY,
                        project_document_path TEXT,
                        logistics_document_ok INTEGER,
                        project_logo_path TEXT,
                        logistics_logo_ok INTEGER,
                        logistics_photos_ok INTEGER,
                        logistics_registration_form_signed_ok INTEGER,
                        logistics_student_consents_signed_ok INTEGER,
                        logistics_cedula_tutor_ok INTEGER,
                        logistics_status TEXT
                    )
                    """
                )
            )
            connection.execute(
                text(
                    """
                    CREATE TABLE project_members (
                        id INTEGER PRIMARY KEY,
                        project_id INTEGER,
                        photo_url TEXT,
                        consent_signed_ok INTEGER,
                        cedula_encargado_ok INTEGER,
                        cedula_estudiante_ok INTEGER
                    )
                    """
                )
            )
            connection.execute(
                text(
                    """
                    INSERT INTO projects VALUES
                        (1, 'document.pdf', 1, 'logo.png', 1, 1, 1, 1, 1, 'pendiente_revision'),
                        (2, 'document.pdf', 1, 'logo.png', 0, 1, 1, 1, 1, 'completo'),
                        (3, 'document.pdf', 1, 'logo.png', 1, 0, 1, 1, 1, 'incompleto')
                    """
                )
            )
            connection.execute(
                text(
                    """
                    INSERT INTO project_members VALUES
                        (1, 1, 'student.jpg', 1, 1, 1),
                        (2, 2, 'student.jpg', 1, 1, 1),
                        (3, 3, 'student.jpg', 1, 1, 1)
                    """
                )
            )

            _reconcile_existing_logistics_statuses(connection)
            rows = connection.execute(
                text("SELECT id, logistics_status, logistics_photos_ok FROM projects")
            )
            statuses = {row.id: row.logistics_status for row in rows}
            photo_flags = {
                row.id: row.logistics_photos_ok
                for row in connection.execute(text("SELECT id, logistics_photos_ok FROM projects"))
            }

        self.assertEqual("completo", statuses[1])
        self.assertEqual("incompleto", statuses[2])
        self.assertEqual("completo", statuses[3])
        self.assertEqual(1, photo_flags[3])
        engine.dispose()

    def test_logistics_template_does_not_offer_resource_validation(self):
        template = Path("app/templates/admin/projects.html").read_text(encoding="utf-8")
        requirements_template = Path("app/templates/admin/requirements.html").read_text(encoding="utf-8")

        self.assertNotIn('name="logistics_requirements_reviewed_ok"', template)
        self.assertIn('name="requirements_internet_ok"', requirements_template)
        self.assertIn('name="requirement_item_confirmed"', requirements_template)
        self.assertIn("Detalle de insumos y materiales", requirements_template)
        self.assertIn("row.getAttribute('action')", requirements_template)
        self.assertNotIn("fetch(row.action", requirements_template)

    def test_projects_can_be_filtered_by_advisor(self):
        template = Path("app/templates/admin/projects.html").read_text(encoding="utf-8")

        self.assertIn('id="projects-filter-advisor"', template)
        self.assertIn("data-project-advisor=", template)
        self.assertIn("matchesAdvisor", template)

    def test_project_summary_reports_completed_and_missing_documents(self):
        complete = Project(
            id=1,
            is_active=True,
            logistics_status="completo",
            project_document_path="document.pdf",
            project_logo_path="logo.png",
            logistics_document_ok=True,
            logistics_logo_ok=True,
            logistics_photos_ok=True,
            logistics_registration_form_signed_ok=True,
            logistics_student_consents_signed_ok=True,
            logistics_cedula_tutor_ok=True,
        )
        pending = Project(id=2, is_active=True, logistics_status="pendiente_revision")
        inactive = Project(id=3, is_active=False, logistics_status="incompleto")

        summary = _build_project_logistics_summary([complete, pending, inactive])

        self.assertEqual(1, summary["completed"])
        self.assertEqual(1, summary["pending"])
        self.assertEqual(1, summary["inactive"])
        self.assertIn("documento digital adjunto", summary["missing_by_project"][2])

    def test_logistics_report_identifies_each_affected_student(self):
        project = Project(
            id=10,
            title="Proyecto de prueba",
            team_name="Equipo ExpoTEC",
            advisor_name="Tutor Ejemplo",
            is_active=True,
            project_document_path="document.pdf",
            project_logo_path="logo.png",
            logistics_document_ok=True,
            logistics_logo_ok=True,
            logistics_photos_ok=False,
            logistics_registration_form_signed_ok=True,
        )
        project.members = [
            ProjectMember(
                full_name="Estudiante Ejemplo",
                section_name="12-1",
                student_number=1,
                photo_url=None,
                consent_signed_ok=True,
            )
        ]

        rows = _build_logistics_pending_report_rows([project], report_type="photo")

        self.assertEqual(1, len(rows))
        self.assertEqual("Fotografía de integrante", rows[0]["pending"])
        self.assertEqual("Estudiante Ejemplo", rows[0]["name"])
        self.assertEqual("12-1", rows[0]["section"])
        self.assertEqual("Proyecto de prueba", rows[0]["project"])
        self.assertEqual("Tutor Ejemplo", rows[0]["tutor"])

    def test_overview_pending_counters_link_reports_center(self):
        template = Path("app/templates/admin/overview.html").read_text(encoding="utf-8")
        controller = Path("app/controllers/admin_controller.py").read_text(encoding="utf-8")

        self.assertIn("reports_page", template)
        self.assertIn("Abrir centro de reportes", template)
        self.assertNotIn(
            'worksheet.auto_filter.ref = f"A5:E{max(worksheet.max_row, 5)}"',
            controller,
        )

    def test_reminder_center_supports_every_audience(self):
        template = Path("app/templates/admin/logistics_reminder.html").read_text(encoding="utf-8")

        self.assertIn('value="students"', template)
        self.assertIn('value="tutors"', template)
        self.assertIn('value="all"', template)
        self.assertIn('name="project_ids"', template)
        self.assertIn("Correo para tutores", template)
        self.assertIn("reminder-batch-progress", template)
        self.assertIn("data-tutor-key", template)
        self.assertIn("tutorGroups", template)
        self.assertIn("formData.set('audience'", template)
        self.assertIn("'Accept': 'application/json'", template)
        self.assertIn("form.getAttribute('action')", template)
        self.assertNotIn("fetch(form.action", template)

        controller = Path("app/controllers/admin_controller.py").read_text(encoding="utf-8")
        self.assertIn('batch_mode = request.form.get("batch_mode") == "1"', controller)
        self.assertIn('return jsonify({"ok": True, "action": action})', controller)
        self.assertIn("await fetch", template)
        self.assertIn('value="save_logo_submission_email"', template)
        self.assertIn('name="logo_submission_email"', template)

    def test_tutor_reminder_contains_group_and_student_pending_items(self):
        from app import create_app

        app = create_app()
        project = Project(
            id=20,
            title="Proyecto con pendientes",
            team_name="Equipo",
            advisor_name="Tutor Ejemplo",
            is_active=True,
            project_logo_path=None,
            logistics_document_ok=False,
            logistics_registration_form_signed_ok=False,
            logistics_cedula_tutor_ok=False,
        )
        project.members = [
            ProjectMember(
                full_name="Estudiante Ejemplo",
                section_name="12-2",
                student_number=1,
                photo_url=None,
                consent_signed_ok=False,
                cedula_encargado_ok=False,
                cedula_estudiante_ok=False,
            )
        ]

        with app.test_request_context("/admin/proyectos/recordatorio"):
            payload = _build_tutor_logistics_reminder_payload(
                project,
                deadline=None,
                institution_name="ExpoTÃ©cnica",
            )

        self.assertIsNotNone(payload)
        self.assertIn("Cédula del tutor", payload["missing_group"])
        self.assertEqual("Estudiante Ejemplo", payload["member_missing"][0]["member"].full_name)
        self.assertIn("Consentimiento informado", payload["member_missing"][0]["items"])

    def test_logo_email_is_included_only_when_logo_must_be_uploaded(self):
        from app import create_app

        app = create_app()
        project = Project(
            id=21,
            title="Proyecto sin logo",
            team_name="Equipo",
            advisor_name="Tutor Ejemplo",
            is_active=True,
            project_logo_path=None,
            logistics_document_ok=True,
            logistics_registration_form_signed_ok=True,
            logistics_cedula_tutor_ok=True,
        )
        project.members = []

        with app.test_request_context("/admin/proyectos/recordatorio"):
            payload = _build_tutor_logistics_reminder_payload(
                project,
                deadline=None,
                institution_name="ExpoTÃ©cnica",
                logo_submission_email="logos@colegio.cr",
            )

        self.assertIn("logos@colegio.cr", payload["plain_body"])
        self.assertIn("mailto:logos@colegio.cr", payload["html_body"])

        project.project_logo_path = "uploads/projects/logo.png"
        project.logistics_logo_ok = False
        with app.test_request_context("/admin/proyectos/recordatorio"):
            payload = _build_tutor_logistics_reminder_payload(
                project,
                deadline=None,
                institution_name="ExpoTÃ©cnica",
                logo_submission_email="logos@colegio.cr",
            )

        self.assertNotIn("logos@colegio.cr", payload["plain_body"])
        self.assertNotIn("mailto:logos@colegio.cr", payload["html_body"])

    def test_tutor_digest_combines_and_orders_all_pending_projects(self):
        from app import create_app

        app = create_app()
        project_b = Project(
            id=22,
            title="Proyecto Zeta",
            advisor_name="Tutor Ejemplo",
            advisor_email="tutor@example.com",
            is_active=True,
            project_logo_path=None,
            logistics_document_ok=False,
            logistics_registration_form_signed_ok=True,
            logistics_cedula_tutor_ok=True,
        )
        project_a = Project(
            id=23,
            title="Proyecto Alfa",
            advisor_name="Tutor Ejemplo",
            advisor_email="tutor@example.com",
            is_active=True,
            project_logo_path="uploads/projects/alfa.png",
            logistics_logo_ok=True,
            logistics_document_ok=True,
            logistics_registration_form_signed_ok=False,
            logistics_cedula_tutor_ok=True,
        )
        project_a.members = []
        project_b.members = []

        with app.test_request_context("/admin/proyectos/recordatorio"):
            payload = _build_tutor_logistics_digest_payload(
                [project_b, project_a],
                deadline=None,
                institution_name="ExpoTÃ©cnica",
                logo_submission_email="logos@colegio.cr",
            )

        self.assertEqual(2, len(payload["project_rows"]))
        self.assertEqual("Proyecto Alfa", payload["project_rows"][0]["project"].title)
        self.assertEqual("Proyecto Zeta", payload["project_rows"][1]["project"].title)
        self.assertIn("2 proyectos", payload["subject"])
        self.assertLess(payload["html_body"].index("Proyecto Alfa"), payload["html_body"].index("Proyecto Zeta"))

        with app.app_context(), patch.object(
            mail_service.SystemSetting,
            "get_value",
            side_effect=lambda key, default=None: default,
        ):
            reminder_data = _build_logistics_reminder_data([project_b, project_a], [])

        self.assertEqual(1, reminder_data["total_tutor_recipients"])
        self.assertEqual([project_b, project_a], reminder_data["tutor_groups"]["tutor@example.com"])

    def test_usher_report_contains_only_confirmed_exposition_assignments(self):
        exposition_judge = Judge(
            id=1,
            full_name="Juez ExposiciÃ³n",
            email="expo@example.com",
            phone="8888-8888",
            role=Judge.ROLE_JUDGE,
            password_hash="test",
            is_active_user=True,
            attendance_confirmed=True,
        )
        documentation_judge = Judge(
            id=2,
            full_name="Juez DocumentaciÃ³n",
            email="doc@example.com",
            role=Judge.ROLE_JUDGE,
            password_hash="test",
            is_active_user=True,
            attendance_confirmed=True,
        )
        project = Project(
            id=30,
            title="Proyecto para exposiciÃ³n",
            team_name="Equipo",
            category="steam",
            is_active=True,
        )
        exposition_assignment = Assignment(
            judge_id=1,
            project_id=30,
            status=Assignment.STATUS_CONFIRMED,
            can_evaluate_documentation=True,
            can_evaluate_exposition=True,
        )
        exposition_assignment.judge = exposition_judge
        documentation_assignment = Assignment(
            judge_id=2,
            project_id=30,
            status=Assignment.STATUS_CONFIRMED,
            can_evaluate_documentation=True,
            can_evaluate_exposition=False,
        )
        documentation_assignment.judge = documentation_judge

        rows = _build_exposition_usher_report_rows(
            {
                "projects": [project],
                "assignments": [exposition_assignment, documentation_assignment],
                "category_map": {"steam": "STEAM"},
            }
        )

        self.assertEqual(1, len(rows))
        self.assertEqual("Juez ExposiciÃ³n", rows[0]["judge"])
        self.assertEqual("Proyecto para exposiciÃ³n", rows[0]["project"])
        self.assertEqual("Sin recinto asignado", rows[0]["location"])

    def test_assignments_page_links_reports_center(self):
        template = Path("app/templates/admin/assignments.html").read_text(encoding="utf-8")

        self.assertIn("reports_page", template)
        self.assertIn("Abrir centro de reportes", template)

    def test_judge_pool_links_reports_center_and_route_keeps_excel(self):
        template = Path("app/templates/admin/judge_pool.html").read_text(encoding="utf-8")
        routes = Path("app/routes/admin_routes.py").read_text(encoding="utf-8")

        self.assertIn("reports_page", template)
        self.assertIn("Abrir reportes", template)
        self.assertIn("/jueces/evaluacion/reporte/excel", routes)

    def test_judge_report_rows_include_attendance_and_assignments(self):
        judge = Judge(
            id=8,
            full_name="JUEZ DE PRUEBA",
            email="JUEZ@EXAMPLE.COM",
            identity="123",
            phone="8888-0000",
            job_title="INGENIERIA",
            institution="EMPRESA ABC",
            role=Judge.ROLE_JUDGE,
            is_active_user=True,
            attendance_confirmed=True,
            needs_parking=True,
            can_evaluate_documentation=True,
            can_evaluate_exposition=True,
            can_evaluate_english=True,
            category_scope="ambas",
        )
        project = Project(
            id=9,
            title="Proyecto Demo",
            team_name="Equipo",
            category="steam",
            is_active=True,
        )
        project.members = [ProjectMember(full_name="Estudiante Ingles", participates_in_english=True)]
        assignment = Assignment(
            id=10,
            judge_id=8,
            project_id=9,
            can_evaluate_documentation=True,
            can_evaluate_exposition=True,
            status=Assignment.STATUS_CONFIRMED,
        )
        assignment.judge = judge
        assignment.project = project

        judge_rows, assignment_rows = _judge_report_rows([judge], [assignment])

        self.assertEqual("Juez de Prueba", judge_rows[0]["name"])
        self.assertEqual("Confirmado", judge_rows[0]["attendance"])
        self.assertEqual("Si", judge_rows[0]["parking"])
        self.assertEqual(1, judge_rows[0]["confirmed_assignments"])
        self.assertEqual(1, judge_rows[0]["english_assignments"])
        self.assertEqual("juez@example.com", assignment_rows[0]["email"])
        self.assertEqual("Si", assignment_rows[0]["document"])
        self.assertEqual("Si", assignment_rows[0]["exposition"])

    def test_judges_report_opens_on_judge_detail_sheet(self):
        source = Path("app/controllers/admin_controller.py").read_text(encoding="utf-8")

        self.assertIn('ws_judges = wb.active', source)
        self.assertIn('ws_judges.title = "Jueces"', source)
        self.assertIn('ws_summary = wb.create_sheet("Resumen")', source)

    def test_logistics_department_does_not_receive_requirements_module(self):
        self.assertNotIn("requirements", ADMIN_DEPARTMENT_MODULE_ACCESS["logistica"])
        self.assertEqual(ACTION_MODULE_MAP["update_project_requirements"], "requirements")


if __name__ == "__main__":
    unittest.main()
